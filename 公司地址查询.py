# -*- coding: utf-8 -*-
import requests
import json
import re
import pandas as pd
import openpyxl


def get_information(company):
    url_base = "https://apis.map.qq.com/jsapi?qt=geoc&key=UGMBZ-CINWR-DDRW5-W52AK-D3ENK-ZEBRC&output=jsonp&pf=jsapi&ref=jsapi&cb=qq.maps._svcb3.geocoder0&addr="
    url = url_base + company

    payload = {}
    headers = {}

    response = requests.request("GET", url, headers=headers, data=payload)

    response_text = findAll(response.text)
    response_json = json.loads(response_text)

    # 公司信息
    company_name = company
    # 省
    province = response_json.get("detail").get("province")
    # 市
    city = response_json.get("detail").get("city")
    # 区
    district = response_json.get("detail").get("district")
    # 街道
    town = response_json.get("detail").get("town")
    # 经度
    pointx = response_json.get("detail").get("pointx")
    # 维度
    pointy = response_json.get("detail").get("pointy")
    # 详细地址
    analysis_address = response_json.get("detail").get("analysis_address")

    company_body = {}
    company_body["company_name"] = company_name
    company_body["province"] = province
    company_body["city"] = city
    company_body["district"] = district
    company_body["town"] = town
    company_body["pointx"] = pointx
    company_body["pointy"] = pointy
    company_body["analysis_address"] = analysis_address

    return company_body


def findAll(s):
    text = s.replace("\n", "")
    pattern = r'\((.*)\)'
    result = re.findall(pattern, text)  
    return result[0]

def add_company_information_to_xlsx(company_information_list):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("公司信息", 0)  
    ws.append(["编号", "标注名称", "地名地址", "经度", "维度", "省", "市", "区", "乡/镇", "详细地址"])  
    for company_information in company_information_list:

        province = company_information.get("province") 
        if(province is None): province = ""
        city = company_information.get("city")
        if(city is None): city = ""
        district = company_information.get("district")
        if(district is None): district = ""
        town = company_information.get("town")
        if(town is None): town = ""
        analysis_address = company_information.get("analysis_address")
        if(analysis_address is None): analysis_address = ""
        address = province + city + district + town
        pointx = company_information.get("pointx")
        if(pointx is None): pointx = ""
        pointy = company_information.get("pointy")
        if(pointy is None): pointy = ""

        ws.cell(row=ws.max_row + 1, column=1, value=ws.max_row)
        ws.cell(row=ws.max_row, column=2, value=company_information.get("company_name"))
        ws.cell(row=ws.max_row, column=3, value=address)
        ws.cell(row=ws.max_row, column=4, value=pointx)  
        ws.cell(row=ws.max_row, column=5, value=pointy)

        ws.cell(row=ws.max_row, column=6, value=province)
        ws.cell(row=ws.max_row, column=7, value=city)
        ws.cell(row=ws.max_row, column=8, value=district)
        ws.cell(row=ws.max_row, column=9, value=town)
        ws.cell(row=ws.max_row, column=10, value=analysis_address)

    wb.save("Company.xlsx")


if __name__ == "__main__":

    with open('Company.txt', 'r', encoding='utf-8') as file:  
        lines = file.readlines()

    # 将每一行转换为列表，不进行任何分割
    company_list = [line.strip() for line in lines]
    # print(company_list)

    company_information_list = []
    for company in company_list:
        company_information_list.append(get_information(company))

    print(company_information_list)

    add_company_information_to_xlsx(company_information_list)

